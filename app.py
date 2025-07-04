# T12 Report Formatter - Modular Architecture
# Handles both Summary and Detail reports with extensibility for future report types

import streamlit as st
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import tempfile
from abc import ABC, abstractmethod

class T12ReportFormatter(ABC):
    """Abstract base class for T12 report formatters"""
    
    def __init__(self, file_path):
        self.wb = load_workbook(file_path)
        self.ws = self.wb.active
        self.header_format = self._detect_header_format()
        
    def format(self):
        """Main formatting method that calls all formatting steps"""
        self._unmerge_cells()
        self._align_header_cells()
        self._set_column_widths()
        self._delete_header_rows()
        self._freeze_panes()
        self._set_row_heights()
        self._hide_gridlines()
        return self._save_output()
    
    def _detect_header_format(self):
        """Detect which header format the file uses"""
        # Check row 3 specifically
        cell_a3 = self.ws["A3"].value
        
        # Standard format has "Location:" in row 3
        # Alternate format has the date in row 3
        if cell_a3 and "Location:" in str(cell_a3):
            return "standard"
        else:
            return "alternate"
    
    def _unmerge_cells(self):
        """Unmerge cells in the header area"""
        for merged_range in list(self.ws.merged_cells.ranges):
            if merged_range.max_row <= 60 and merged_range.max_col <= 14:
                self.ws.unmerge_cells(str(merged_range))
    
    def _align_header_cells(self):
        """Align left cells based on format"""
        if self.header_format == "standard":
            # Standard format: align A6-A8
            for row in range(6, 9):
                self.ws[f"A{row}"].alignment = Alignment(horizontal='left')
        else:
            # Alternate format: align A1-A3
            for row in range(1, 4):
                self.ws[f"A{row}"].alignment = Alignment(horizontal='left')
    
    def _set_column_widths(self):
        """Set Column Widths B–N to 12 pixels"""
        for col in range(2, 15):
            self.ws.column_dimensions[get_column_letter(col)].width = 12
    
    def _delete_header_rows(self):
        """Delete header rows based on detected format"""
        if self.header_format == "standard":
            # Standard format: delete rows 1,2,3,4,5,9,10,11
            # This removes the 5 top rows and the "Reporting Book/As of Date" rows
            for row in sorted([1, 2, 3, 4, 5, 9, 10, 11], reverse=True):
                self.ws.delete_rows(row)
        else:
            # Alternate format: need to match standard output structure
            # Delete rows 4,5,6 (Reporting Book, As of Date, Location)
            # This leaves 3 blank rows (7,8,9) + 1 more (10) = 4 rows before Income
            # But standard has only 3 blank rows before Income, so delete one more
            for row in sorted([4, 5, 6, 7], reverse=True):
                self.ws.delete_rows(row)
        
        # Delete "Created on:" footer row if it exists
        for row in range(self.ws.max_row, max(1, self.ws.max_row - 10), -1):
            cell_value = self.ws[f"A{row}"].value
            if cell_value and "Created on:" in str(cell_value):
                self.ws.delete_rows(row)
                break
    
    def _freeze_panes(self):
        """Freeze pane at B6 (after row deletion)"""
        # Ensure we're setting it correctly for openpyxl
        self.ws.freeze_panes = self.ws['B6']
    
    def _set_row_heights(self):
        """Set row heights for header rows"""
        self.ws.row_dimensions[1].height = 18
        self.ws.row_dimensions[2].height = 18
        self.ws.row_dimensions[3].height = 16
    
    def _hide_gridlines(self):
        """Hide gridlines"""
        self.ws.sheet_view.showGridLines = False
    
    @abstractmethod
    def _get_report_type_suffix(self):
        """Return the suffix for the filename - must be implemented by subclasses"""
        pass
    
    def _save_output(self):
        """Save the formatted file with appropriate naming"""
        # Get date from A3 (which becomes A3 after deletions)
        date_value = self.ws["A3"].value if self.ws["A3"].value else "NoDate"
        try:
            parsed_date = datetime.strptime(str(date_value), "%B %d, %Y")
            formatted_date = parsed_date.strftime("%Y-%m")
        except ValueError:
            try:
                # Try alternative format
                parsed_date = datetime.strptime(str(date_value), "%m/%d/%Y")
                formatted_date = parsed_date.strftime("%Y-%m")
            except ValueError:
                formatted_date = "Unknown_Date"
        
        # Get property name from A1
        property_name = self.ws["A1"].value or "Property"
        safe_property = str(property_name).replace(" ", "_").replace("/", "-").strip()
        
        # Create filename
        suffix = self._get_report_type_suffix()
        filename = f"{safe_property}_{suffix}_{formatted_date}.xlsx"
        output_path = os.path.join(tempfile.gettempdir(), filename)
        self.wb.save(output_path)
        
        return output_path

class T12SummaryFormatter(T12ReportFormatter):
    """Formatter for T12 Summary reports"""
    
    def _delete_header_rows(self):
        """Delete header rows including row 60 if it exists"""
        super()._delete_header_rows()
        
        # After deleting 4 rows, original row 60 becomes row 56
        if self.ws.max_row >= 56:
            if not any(self.ws.cell(row=56, column=col).value for col in range(1, 15)):
                self.ws.delete_rows(56)
    
    def _get_report_type_suffix(self):
        return "T12 Summary"

class T12DetailFormatter(T12ReportFormatter):
    """Formatter for T12 Detail reports"""
    
    def _unmerge_cells(self):
        """Unmerge ALL cells in the entire worksheet for detail reports"""
        # Detail reports have merged cells throughout, not just in header
        for merged_range in list(self.ws.merged_cells.ranges):
            self.ws.unmerge_cells(str(merged_range))
    
    def _get_report_type_suffix(self):
        return "T12 Income Statement"

def format_report(file_path, report_type):
    """Factory function to create appropriate formatter"""
    if report_type == "summary":
        formatter = T12SummaryFormatter(file_path)
    elif report_type == "detail":
        formatter = T12DetailFormatter(file_path)
    else:
        raise ValueError(f"Unknown report type: {report_type}")
    
    return formatter.format()

# Streamlit UI
st.set_page_config(page_title="T12 Report Formatter", page_icon="📊", layout="wide")

st.title("📊 T12 Report Formatter")
st.write("Format T12 Summary and Detail reports for multifamily properties")

# Create two columns for the interface
col1, col2 = st.columns([1, 2])

with col1:
    st.subheader("Upload & Configure")
    
    # Report type selection
    report_type = st.selectbox(
        "Select Report Type",
        ["summary", "detail"],
        format_func=lambda x: "T12 Summary" if x == "summary" else "T12 Income Statement (Detail)"
    )
    
    # File upload
    uploaded_file = st.file_uploader(
        f"Upload your T12 {report_type.title()} Excel file", 
        type=["xlsx"]
    )

with col2:
    st.subheader("Formatting Steps")
    
    if report_type == "summary":
        st.markdown("""
        **Summary Report Formatting:**
        1. ✓ Unmerge header cells
        2. ✓ Align property info left (A1-A3)
        3. ✓ Set column widths to 12px
        4. ✓ Delete rows 4-6, 10 & row 60
        5. ✓ Delete "Created on:" footer row
        6. ✓ Freeze pane at B6
        7. ✓ Set row heights
        8. ✓ Hide gridlines
        9. ✓ Save as: `PropertyName_T12 Summary_YYYY-MM.xlsx`
        """)
    else:
        st.markdown("""
        **Detail Report Formatting:**
        1. ✓ Unmerge ALL cells (entire document)
        2. ✓ Align property info left (A1-A3)
        3. ✓ Set column widths to 12px
        4. ✓ Delete rows 4-6, 10
        5. ✓ Delete "Created on:" footer row
        6. ✓ Freeze pane at B6
        7. ✓ Set row heights
        8. ✓ Hide gridlines
        9. ✓ Save as: `PropertyName_T12 Income Statement_YYYY-MM.xlsx`
        """)

# Process file if uploaded
if uploaded_file:
    try:
        with st.spinner(f"Formatting {report_type} report..."):
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                tmp.write(uploaded_file.read())
                formatted_path = format_report(tmp.name, report_type)
        
        st.success("✅ Report formatted successfully!")
        
        with open(formatted_path, "rb") as f:
            st.download_button(
                label=f"📥 Download Formatted T12 {report_type.title()}",
                data=f,
                file_name=os.path.basename(formatted_path),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
        # Cleanup
        os.unlink(tmp.name)
        
    except Exception as e:
        st.error(f"❌ Error formatting report: {str(e)}")
        st.exception(e)

# Footer with instructions
st.markdown("---")
st.markdown("""
### 🚀 Coming Soon
- Additional report types (Rent Roll, Operating Statement, etc.)
- Batch processing for multiple files
- Custom formatting templates
- API for integration with other CRE tools
""")
